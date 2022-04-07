from docx import Document
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.utils.dataframe import dataframe_to_rows
import time

from openpyxl.styles import Font
from openpyxl.styles import Alignment
import os

name_spo = 'БРИТ'
path = 'data'

# Создаем список папок
lst_dir = []
for value in os.listdir(path):
    path_value = os.path.join(path, value)
    if os.path.isdir(path_value):
        lst_dir.append(value)
# Создаем документ Excel
wb = openpyxl.Workbook()
# Переименовываем лист чтобы в итоговом файле не было пустого листа
ren_sheet = wb['Sheet']
ren_sheet.title = 'Сводная таблица'
# Создаем заголовки для сводной таблицы
wb['Сводная таблица']['A1'] = 'Наименование ПОО'
wb['Сводная таблица']['B1'] = 'Наименование государственной услуги'
wb['Сводная таблица']['C1'] = 'Категория потребителей государственной услуги'
wb['Сводная таблица']['D1'] = 'Уникальный номер реестровой записи'
wb['Сводная таблица']['E1'] = 'Профессии и специальности по программам среднего профессионального образования'
wb['Сводная таблица']['F1'] = 'Категория потребителей'
wb['Сводная таблица']['G1'] = 'Формы образования и формы реализации образовательных программ'
wb['Сводная таблица']['H1'] = 'Госзадание №1'
wb['Сводная таблица']['I1'] = 'Госзадание №2'
wb['Сводная таблица']['J1'] = 'исполнено на отчетную дату'
wb['Сводная таблица']['K1'] = 'причина отклонения'
wb['Сводная таблица']['L1'] = '5% Допустимое отклонение в единицах'
wb['Сводная таблица']['M1'] = 'Реальное отклонение в единицах'
wb['Сводная таблица']['N1'] = 'Отклонение в процентах'
wb['Сводная таблица']['O1'] = 'Отклонение больше 5%'

wb['Сводная таблица'].column_dimensions['B'].width = 50
wb['Сводная таблица'].column_dimensions['C'].width = 50
wb['Сводная таблица'].column_dimensions['D'].width = 30
wb['Сводная таблица'].column_dimensions['E'].width = 70
wb['Сводная таблица'].column_dimensions['F'].width = 50
wb['Сводная таблица'].column_dimensions['G'].width = 15
wb['Сводная таблица'].column_dimensions['H'].width = 15
wb['Сводная таблица'].column_dimensions['I'].width = 15
wb['Сводная таблица'].column_dimensions['J'].width = 15
wb['Сводная таблица'].column_dimensions['K'].width = 60
wb['Сводная таблица'].column_dimensions['L'].width = 30
wb['Сводная таблица'].column_dimensions['M'].width = 30

# Создаем заливку которой будет отмечать ячейки где отклонение больше 5 %
redFill = PatternFill(start_color='960018',
                                 end_color='960018',
                                 fill_type='solid')
# Создаем новые листы по количеству папок
print(lst_dir)
for i in range(len(lst_dir)):
    wb.create_sheet(title=f'{lst_dir[i]}', index=i + 1)
print(wb.sheetnames)

# перебираем папки
for dir in os.listdir(path):
    if os.path.isdir(f'{path}/{dir}'):
        # Получаем путь к файлу
        path_to_file = f'{path}/{dir}/'
        # Открываем файл Госзадания №1
        check_name_file = ''

        try:
            check_name_file = f'{dir} Госзадание №1'
            # df_1 = pd.read_excel(f'{path_to_file}Госзадание №1.xlsx')
            df_1 = pd.read_excel(f'{path_to_file}{dir} Госзадание №1.xlsx')
            # Отбираем колонки чтобы сделать из них мультииндекс
            ml_1 = df_1[['Наименование государственной услуги', 'Категория потребителей государственной услуги',
                         'Уникальный номер реестровой записи',
                         'Профессии и специальности по программам среднего профессионального образования']]

            # Отбираем колонки с данными
            data_1 = df_1[
                ['Категория потребителей', 'Формы образования и формы реализации образовательных программ',
                 'Госзадание №1']]

            # Создаем мультиндекс
            ml_df1 = pd.MultiIndex.from_frame(ml_1, names=['Наименование государственной услуги',
                                                           'Категория потребителей государственной услуги',
                                                           'Уникальный номер реестровой записи',
                                                           'Профессии и специальности по программам среднего профессионального образования'])

            # Присваиваем мультиндекс
            data_1.index = ml_df1

            # Открываем госзадание №2
            check_name_file = f'{dir} Госзадание №2'
            df_2 = pd.read_excel(f'{path_to_file}{dir} Госзадание №2.xlsx')

            # Создаем мультииндекс
            ml_2 = df_2[['Наименование государственной услуги', 'Категория потребителей государственной услуги',
                         'Уникальный номер реестровой записи',
                         'Профессии и специальности по программам среднего профессионального образования']]

            # Забираем данные которые нам нужны
            data_2 = df_2[['Госзадание №2']]

            ml_df2 = pd.MultiIndex.from_frame(ml_2, names=['Наименование государственной услуги',
                                                           'Категория потребителей государственной услуги',
                                                           'Уникальный номер реестровой записи',
                                                           'Профессии и специальности по программам среднего профессионального образования'])

            data_2.index = ml_df2

            gos_zad_df = pd.concat([data_1, data_2], axis=1, join='outer')

            # Открываем файл отчета
            check_name_file = f'{dir} Отчет'
            gos_report = pd.read_excel(f'{path_to_file}{dir} Госзадание Отчет.xlsx')

            # Создаем мультииндекс
            ml_report_df = gos_report[
                ['Наименование государственной услуги', 'Категория потребителей государственной услуги',
                 'Уникальный номер реестровой записи',
                 'Профессии и специальности по программам среднего профессионального образования']]

            data_report = gos_report[['исполнено на отчетную дату', 'причина отклонения']]

            ml_report = pd.MultiIndex.from_frame(ml_report_df, names=['Наименование государственной услуги',
                                                                      'Категория потребителей государственной услуги',
                                                                      'Уникальный номер реестровой записи',
                                                                      'Профессии и специальности по программам среднего профессионального образования'])

            data_report.index = ml_report

            # Соединяем в итоговый датафрейм
            itog_df = pd.concat([gos_zad_df, data_report], axis=1, join='outer')



            # Добавляем нужные столбы
            itog_df['5% Допустимое отклонение в единицах'] = (itog_df['Госзадание №2'] / 100 * 5).round(0)

            itog_df['Реальное отклонение в единицах'] = (
                    itog_df['Госзадание №2'] - itog_df['исполнено на отчетную дату']).abs()

            itog_df['Отклонение в процентах'] = (
                    itog_df['Реальное отклонение в единицах'] * 100 / itog_df['Госзадание №2']).round(2)

            itog_df['Отклонение больше 5%'] = itog_df['Отклонение в процентах'].apply(
                lambda x: 'Да' if x > 5 else 'Нет')

            itog_df['Отклонение в процентах'] = itog_df['Отклонение в процентах'].astype(str) + '%'

            itog_df = itog_df.reset_index()

            itog_df.insert(0, 'Наименование ПОО', dir)
            for r in dataframe_to_rows(itog_df, index=False, header=False):
                if len(r) != 1:
                    wb[f'Сводная таблица'].append(r)

            for r in dataframe_to_rows(itog_df, index=False, header=True):
                if len(r) != 1:
                    wb[f'{dir}'].append(r)
            # Устанавливаем цвет ячейки
            for cell in wb[f'{dir}']['O']:
                if cell.value == 'Да':
                    cell.fill = redFill
            # Устанавливаем размер колонок
            wb[f'{dir}'].column_dimensions['B'].width = 50
            wb[f'{dir}'].column_dimensions['C'].width = 50
            wb[f'{dir}'].column_dimensions['D'].width = 30
            wb[f'{dir}'].column_dimensions['E'].width = 70
            wb[f'{dir}'].column_dimensions['F'].width = 50
            wb[f'{dir}'].column_dimensions['G'].width = 15
            wb[f'{dir}'].column_dimensions['H'].width = 15
            wb[f'{dir}'].column_dimensions['I'].width = 15
            wb[f'{dir}'].column_dimensions['J'].width = 15
            wb[f'{dir}'].column_dimensions['K'].width = 60
            wb[f'{dir}'].column_dimensions['L'].width = 30
            wb[f'{dir}'].column_dimensions['M'].width = 30


        except FileNotFoundError:
            print(f'В папке {dir} Файл {check_name_file} Не найден!!! Проверьте названия файлов внутри папки')
        except KeyError:
            print('Проверьте названия колонок в файле')
        else:
            # # Добавляем перенос
            # for cell in wb['Сводная таблица']['B']:
            #     cell.alignment = Alignment(wrap_text=True)
            # for cell in wb['Сводная таблица']['C']:
            #     cell.alignment = Alignment(wrap_text=True)

            # Выделяем ячейку в зависимости от значения
            for cell in wb['Сводная таблица']['O']:
                if cell.value == 'Да':
                    cell.fill = redFill

            # Получаем текущее время для того чтобы использовать в названии
            t = time.localtime()
            current_time = time.strftime('%H_%M_%S', t)
            # Сохраняем итоговый файл
            wb.save(f'data/Сводный отчет по выполнению госзадания.xlsx {current_time}.xlsx')
