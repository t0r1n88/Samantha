import pandas as pd
import  openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import os
from docxtpl import DocxTemplate
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import datetime
from datetime import date
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart, Reference, PieChart, PieChart3D, Series
import warnings
from openpyxl.styles import Font
from openpyxl.styles import Alignment


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def select_file_data_calculate_budget_spo():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global name_file_data_calculate_budget_spo
    # Получаем путь к файлу
    name_file_data_calculate_budget_spo = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_end_folder_calculate_budget_spo():
    """
    Функция для выбора папки куда будет генерироваться файл
    :return:
    """
    global path_to_end_folder_calculate_budget_spo
    path_to_end_folder_calculate_budget_spo = filedialog.askdirectory()

def calculate_budget_spo():
    # получаем список листов в таблице
    try:
        sheetnames = openpyxl.load_workbook(name_file_data_calculate_budget_spo).sheetnames

        # Создаем документ openpyxl

        wb = openpyxl.Workbook()
        wb.create_sheet(title='Сводная таблица', index=0)
        wb.create_sheet(title='ГАПОУ', index=1)
        wb.create_sheet(title='ГБПОУ', index=2)

        # получаем базовый датафрейм
        base_df = pd.read_excel(name_file_data_calculate_budget_spo, sheet_name='СВОД')

        # Итерируемся открываем листы
        for sheet in sheetnames:
            if sheet != 'СВОД':
                # Считываем датафрейм
                temp_df = pd.read_excel(name_file_data_calculate_budget_spo, sheet_name=sheet)
                base_df = pd.concat([base_df, temp_df], axis=0, ignore_index=True)

        clean_df = clean_data(base_df)

        #
        # Разделяем по ГАПОУ и ГБПОУ
        g_df = clean_df.copy()

        group_df = g_df.groupby(['Наименование учреждения'], as_index=False).agg({'Сумма, тыс. рублей': 'sum'})

        group_df['Тип'] = group_df['Наименование учреждения'].apply(lambda x: 'ГАПОУ' if 'ГАПОУ' in x else 'ГБПОУ')

        ga_df = group_df[group_df['Тип'] == 'ГАПОУ']
        gb_df = group_df[group_df['Тип'] == 'ГБПОУ']

        # Удаляем колонки  с типом
        out_ga_df = ga_df.drop(columns='Тип')
        out_gb_df = gb_df.drop(columns='Тип')

        # Записываем результаты в соответствующие листы
        for r in dataframe_to_rows(clean_df, index=False, header=True):
            if len(r) != 1:
                wb['Сводная таблица'].append(r)

        for r in dataframe_to_rows(out_ga_df, index=False, header=True):
            if len(r) != 1:
                wb['ГАПОУ'].append(r)

        for r in dataframe_to_rows(out_gb_df, index=False, header=True):
            if len(r) != 1:
                wb['ГБПОУ'].append(r)

        # Устанавливаем ширину колонок
        wb['Сводная таблица'].column_dimensions['A'].width = 5
        wb['Сводная таблица'].column_dimensions['B'].width = 50

        wb['Сводная таблица'].column_dimensions['C'].width = 50
        wb['Сводная таблица'].column_dimensions['D'].width = 20
        wb['Сводная таблица'].column_dimensions['E'].width = 50
        wb['Сводная таблица'].column_dimensions['F'].width = 50

        wb['ГАПОУ'].column_dimensions['A'].width = 50
        wb['ГБПОУ'].column_dimensions['A'].width = 50

        # Получаем текущее время для того чтобы использовать в названии
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Сохраняем итоговый файл
        wb.save(f'{path_to_end_folder_calculate_budget_spo}/Бюджетная заявка от СПО.xlsx {current_time}.xlsx')
    except NameError:
        messagebox.showinfo('ЦОПП Бурятия', f'Выберите файл с данными и папку куда будут генерироваться файлы')

    else:
        messagebox.showinfo('ЦОПП Бурятия','Подсчет успешно завершен!')

def clean_data(raw_df:pd.DataFrame) ->pd.DataFrame:
    raw_df['Сумма, тыс. рублей'].fillna('0.0', inplace=True)
    # base_df.to_excel('Общий.xlsx',index=False)
    # Очищаем столбец от пробелов
    raw_df['Сумма, тыс. рублей'] = raw_df['Сумма, тыс. рублей'].astype('str')

    raw_df['Сумма, тыс. рублей'] = raw_df['Сумма, тыс. рублей'].apply(lambda x: x.replace(' ', ''))
    # Меняем запятую на точку
    raw_df['Сумма, тыс. рублей'] = raw_df['Сумма, тыс. рублей'].apply(lambda x: x.replace(',', '.'))

    raw_df['Сумма, тыс. рублей'] = raw_df['Сумма, тыс. рублей'].astype('float')
    return raw_df


if __name__=='__main__':
    window = Tk()
    window.title('ЦОПП Бурятия')
    window.geometry('700x860')
    window.resizable(False, False)


    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку подсчета бюджетных заявок
    tab_calculate_budget_spo = ttk.Frame(tab_control)
    tab_control.add(tab_calculate_budget_spo, text='Бюджетные заявки СПО')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку Бюджетных заявок СПО
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_calculate_budget_spo,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nПодсчет бюджетных заявок СПО')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_calculate_budget_spo = resource_path('logo.png')
    img_calculate_budget_spo = PhotoImage(file=path_to_calculate_budget_spo)
    Label(tab_calculate_budget_spo,
          image=img_calculate_budget_spo
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_calculate_budget_spo = LabelFrame(tab_calculate_budget_spo, text='Подготовка')
    frame_data_for_calculate_budget_spo.grid(column=0, row=2, padx=10)


    # Создаем кнопку Выбрать файл с данными
    btn_data_doc = Button(frame_data_for_calculate_budget_spo, text='1) Выберите таблицу с данными', font=('Arial Bold', 20),
                          command=select_file_data_calculate_budget_spo
                          )
    btn_data_doc.grid(column=0, row=3, padx=10, pady=10)
    #
    # Создаем кнопку для выбора папки куда будет генерироваться файл

    btn_choose_end_folder_doc = Button(frame_data_for_calculate_budget_spo, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                       command=select_end_folder_calculate_budget_spo
                                       )
    btn_choose_end_folder_doc.grid(column=0, row=4, padx=10, pady=10)

    # Создаем кнопку для создания документов из таблиц с произвольной структурой
    btn_create_files_other = Button(tab_calculate_budget_spo, text='Подсчитать',
                                    font=('Arial Bold', 20),
                                    command=calculate_budget_spo
                                    )
    btn_create_files_other.grid(column=0, row=5, padx=10, pady=10)
    window.mainloop()