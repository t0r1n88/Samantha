import pandas as pd
import  openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import time

def clean_data(raw_df:pd.DataFrame) ->pd.DataFrame:
    raw_df['Сумма, тыс. рублей'].fillna('0.0', inplace=True)
    # base_df.to_excel('Общий.xlsx',index=False)
    # Очищаем столбец от пробелов
    raw_df['Сумма, тыс. рублей'] = raw_df['Сумма, тыс. рублей'].astype('str')

    raw_df['Сумма, тыс. рублей'] = raw_df['Сумма, тыс. рублей'].apply(lambda x: x.replace(' ', ''))
    # Меняем запятую на точку
    raw_df['Сумма, тыс. рублей'] = raw_df['Сумма, тыс. рублей'].apply(lambda x: x.replace(',', '.'))

    raw_df['Сумма, тыс. рублей'] = raw_df['Сумма, тыс. рублей'].astype('float')
    return raw_df


path_to_end_folder_calculate_budget_spo = 'data'
name_file_data_calculate_budget_spo = 'data/Бюджетная заявка от отдела СПО.xlsx'
# получаем список листов в таблице
sheetnames = openpyxl.load_workbook(name_file_data_calculate_budget_spo).sheetnames

# Создаем документ openpyxl

wb = openpyxl.Workbook()
wb.create_sheet(title='Сводная таблица',index=0)
wb.create_sheet(title='ГАПОУ',index=1)
wb.create_sheet(title='ГБПОУ',index=2)

# получаем базовый датафрейм
base_df = pd.read_excel(name_file_data_calculate_budget_spo, sheet_name='СВОД')

#Итерируемся открываем листы
for sheet in sheetnames:
    if sheet !='СВОД':
        # Считываем датафрейм
        temp_df = pd.read_excel(name_file_data_calculate_budget_spo, sheet_name=sheet)
        base_df= pd.concat([base_df,temp_df],axis=0,ignore_index=True)


clean_df = clean_data(base_df)




#
#Разделяем по ГАПОУ и ГБПОУ
g_df = clean_df.copy()

group_df = g_df.groupby(['Наименование учреждения'],as_index=False).agg({'Сумма, тыс. рублей':'sum'})


group_df['Тип'] = group_df['Наименование учреждения'].apply(lambda x:'ГАПОУ' if 'ГАПОУ' in x else 'ГБПОУ')


ga_df = group_df[group_df['Тип'] == 'ГАПОУ']
gb_df = group_df[group_df['Тип'] == 'ГБПОУ']

# Удаляем колонки  с типом
out_ga_df=ga_df.drop(columns='Тип')
out_gb_df = gb_df.drop(columns='Тип')



# Записываем результаты в соответствующие листы
for r in dataframe_to_rows(clean_df,index=False,header=True):
    if len(r) != 1:
        wb['Сводная таблица'].append(r)

for r in dataframe_to_rows(out_ga_df,index=False,header=True):
    if len(r) != 1:
        wb['ГАПОУ'].append(r)

for r in dataframe_to_rows(out_gb_df,index=False,header=True):
    if len(r) != 1:
        wb['ГБПОУ'].append(r)

 # Получаем текущее время для того чтобы использовать в названии
t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)
# Сохраняем итоговый файл
wb.save(f'{path_to_end_folder_calculate_budget_spo}/Бюджетная заявка от СПО.xlsx {current_time}.xlsx')
