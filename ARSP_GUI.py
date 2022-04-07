from docx import Document
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def select_folder_data():
    """
    Функция для выбора папки где находятся данные
    :return:
    """
    global path_to_data
    path_to_data = filedialog.askdirectory()

def select_end_folder():
    """
    Функция для выбора папки где находятся данные
    :return:
    """
    global path_to_end_folder
    path_to_end_folder = filedialog.askdirectory()

def proccessing_scope_state_task():
    """
    Функция для подсчета и обработки данных госзаданий
    :return:
    """
    # Создаем переменную для типа создаваемого документа
    status_rb_type_doc = group_rb_type_doc.get()
    # если статус == 0 то создаем ищем папки внутри выбранной папки и перебираем их
    if status_rb_type_doc == 0:
        # Создаем список папок
        lst_dir = []
        for value in os.listdir(path_to_data):
            path_value = os.path.join(path_to_data, value)
            if os.path.isdir(path_value):
                lst_dir.append(value)
        # Создаем документ Excel
        wb = openpyxl.Workbook()
        # Переименовываем лист чтобы в итоговом файле не было пустого листа
        ren_sheet = wb['Sheet']
        ren_sheet.title = 'Сводная таблица'
        # Создаем заголовки для сводной таблицы
        wb['Сводная таблица']['A1'] = 'Наименование ПОО'
        wb['Сводная таблица']['B1'] = 'Наименование государственной услуги'
        wb['Сводная таблица']['C1'] = 'Категория потребителей государственной услуги'
        wb['Сводная таблица']['D1'] = 'Уникальный номер реестровой записи'
        wb['Сводная таблица']['E1'] = 'Профессии и специальности по программам среднего профессионального образования'
        wb['Сводная таблица']['F1'] = 'Категория потребителей'
        wb['Сводная таблица']['G1'] = 'Формы образования и формы реализации образовательных программ'
        wb['Сводная таблица']['H1'] = 'Госзадание №1'
        wb['Сводная таблица']['I1'] = 'Госзадание №2'
        wb['Сводная таблица']['J1'] = 'исполнено на отчетную дату'
        wb['Сводная таблица']['K1'] = 'причина отклонения'
        wb['Сводная таблица']['L1'] = '5% Допустимое отклонение в единицах'
        wb['Сводная таблица']['M1'] = 'Реальное отклонение в единицах'
        wb['Сводная таблица']['N1'] = 'Отклонение в процентах'
        wb['Сводная таблица']['O1'] = 'Отклонение больше 5%'

        wb['Сводная таблица'].column_dimensions['B'].width = 50
        wb['Сводная таблица'].column_dimensions['C'].width = 50
        wb['Сводная таблица'].column_dimensions['D'].width = 30
        wb['Сводная таблица'].column_dimensions['E'].width = 70
        wb['Сводная таблица'].column_dimensions['F'].width = 50
        wb['Сводная таблица'].column_dimensions['G'].width = 15
        wb['Сводная таблица'].column_dimensions['H'].width = 15
        wb['Сводная таблица'].column_dimensions['I'].width = 15
        wb['Сводная таблица'].column_dimensions['J'].width = 15
        wb['Сводная таблица'].column_dimensions['K'].width = 60
        wb['Сводная таблица'].column_dimensions['L'].width = 30
        wb['Сводная таблица'].column_dimensions['M'].width = 30

        # Создаем заливку которой будет отмечать ячейки где отклонение больше 5 %
        redFill = PatternFill(start_color='960018',
                              end_color='960018',
                              fill_type='solid')
        # Создаем новые листы по количеству папок
        for i in range(len(lst_dir)):
            wb.create_sheet(title=f'{lst_dir[i]}', index=i + 1)

        # перебираем папки
        for name_dir in os.listdir(path_to_data):
            if os.path.isdir(f'{path_to_data}/{name_dir}'):
                # Получаем путь к файлу
                path_to_file = f'{path_to_data}/{name_dir}/'
                # Открываем файл Госзадания №1
                check_name_file = ''

                try:
                    check_name_file = f'{name_dir} Госзадание №1'
                    df_1 = pd.read_excel(f'{path_to_file}{name_dir} Госзадание №1.xlsx')
                    # Отбираем колонки чтобы сделать из них мультииндекс
                    ml_1 = df_1[['Наименование государственной услуги', 'Категория потребителей государственной услуги',
                                 'Уникальный номер реестровой записи',
                                 'Профессии и специальности по программам среднего профессионального образования']]

                    # Отбираем колонки с данными
                    data_1 = df_1[
                        ['Категория потребителей', 'Формы образования и формы реализации образовательных программ',
                         'Госзадание №1']]

                    # Создаем мультиндекс
                    ml_df1 = pd.MultiIndex.from_frame(ml_1, names=['Наименование государственной услуги',
                                                                   'Категория потребителей государственной услуги',
                                                                   'Уникальный номер реестровой записи',
                                                                   'Профессии и специальности по программам среднего профессионального образования'])

                    # Присваиваем мультиндекс
                    data_1.index = ml_df1

                    # Открываем госзадание №2
                    check_name_file = f'{name_dir} Госзадание №2'
                    df_2 = pd.read_excel(f'{path_to_file}{name_dir} Госзадание №2.xlsx')

                    # Создаем мультииндекс
                    ml_2 = df_2[['Наименование государственной услуги', 'Категория потребителей государственной услуги',
                                 'Уникальный номер реестровой записи',
                                 'Профессии и специальности по программам среднего профессионального образования']]

                    # Забираем данные которые нам нужны
                    data_2 = df_2[['Госзадание №2']]

                    ml_df2 = pd.MultiIndex.from_frame(ml_2, names=['Наименование государственной услуги',
                                                                   'Категория потребителей государственной услуги',
                                                                   'Уникальный номер реестровой записи',
                                                                   'Профессии и специальности по программам среднего профессионального образования'])

                    data_2.index = ml_df2

                    gos_zad_df = pd.concat([data_1, data_2], axis=1, join='outer')

                    # Открываем файл отчета
                    check_name_file = f'{name_dir} Отчет'
                    gos_report = pd.read_excel(f'{path_to_file}{name_dir} Госзадание Отчет.xlsx')

                    # Создаем мультииндекс
                    ml_report_df = gos_report[
                        ['Наименование государственной услуги', 'Категория потребителей государственной услуги',
                         'Уникальный номер реестровой записи',
                         'Профессии и специальности по программам среднего профессионального образования']]

                    data_report = gos_report[['исполнено на отчетную дату', 'причина отклонения']]

                    ml_report = pd.MultiIndex.from_frame(ml_report_df, names=['Наименование государственной услуги',
                                                                              'Категория потребителей государственной услуги',
                                                                              'Уникальный номер реестровой записи',
                                                                              'Профессии и специальности по программам среднего профессионального образования'])

                    data_report.index = ml_report

                    # Соединяем в итоговый датафрейм
                    itog_df = pd.concat([gos_zad_df, data_report], axis=1, join='outer')

                    # Добавляем нужные столбы
                    itog_df['5% Допустимое отклонение в единицах'] = (itog_df['Госзадание №2'] / 100 * 5).round(0)

                    itog_df['Реальное отклонение в единицах'] = (
                            itog_df['Госзадание №2'] - itog_df['исполнено на отчетную дату']).abs()

                    itog_df['Отклонение в процентах'] = (
                            itog_df['Реальное отклонение в единицах'] * 100 / itog_df['Госзадание №2']).round(2)

                    itog_df['Отклонение больше 5%'] = itog_df['Отклонение в процентах'].apply(
                        lambda x: 'Да' if x > 5 else 'Нет')

                    itog_df['Отклонение в процентах'] = itog_df['Отклонение в процентах'].astype(str) + '%'

                    itog_df = itog_df.reset_index()

                    itog_df.insert(0, 'Наименование ПОО', name_dir)
                    for r in dataframe_to_rows(itog_df, index=False, header=False):
                        if len(r) != 1:
                            wb[f'Сводная таблица'].append(r)

                    for r in dataframe_to_rows(itog_df, index=False, header=True):
                        if len(r) != 1:
                            wb[f'{name_dir}'].append(r)
                    # Устанавливаем цвет ячейки
                    for cell in wb[f'{name_dir}']['O']:
                        if cell.value == 'Да':
                            cell.fill = redFill
                    # Устанавливаем размер колонок
                    wb[f'{name_dir}'].column_dimensions['B'].width = 50
                    wb[f'{name_dir}'].column_dimensions['C'].width = 50
                    wb[f'{name_dir}'].column_dimensions['D'].width = 30
                    wb[f'{name_dir}'].column_dimensions['E'].width = 70
                    wb[f'{name_dir}'].column_dimensions['F'].width = 50
                    wb[f'{name_dir}'].column_dimensions['G'].width = 15
                    wb[f'{name_dir}'].column_dimensions['H'].width = 15
                    wb[f'{name_dir}'].column_dimensions['I'].width = 15
                    wb[f'{name_dir}'].column_dimensions['J'].width = 15
                    wb[f'{name_dir}'].column_dimensions['K'].width = 60
                    wb[f'{name_dir}'].column_dimensions['L'].width = 30
                    wb[f'{name_dir}'].column_dimensions['M'].width = 30

                except FileNotFoundError:
                    print(f'В папке {name_dir} Файл {check_name_file} Не найден!!! Проверьте названия файлов внутри папки')
                except KeyError:
                    print(f'Проверьте названия колонок в файле {check_name_file}')
                else:
                    # # Добавляем перенос
                    # for cell in wb['Сводная таблица']['B']:
                    #     cell.alignment = Alignment(wrap_text=True)
                    # for cell in wb['Сводная таблица']['C']:
                    #     cell.alignment = Alignment(wrap_text=True)

                    # Выделяем ячейку в зависимости от значения
                    for cell in wb['Сводная таблица']['O']:
                        if cell.value == 'Да':
                            cell.fill = redFill

        # Получаем текущее время для того чтобы использовать в названии
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Сохраняем итоговый файл
        wb.save(f'{path_to_end_folder}/Сводный отчет по выполнению госзадания {current_time}.xlsx')
        messagebox.showinfo('ЦОПП Бурятия', 'Обработка успешно завершена!!!')
    else:
        name_poo = path_to_data.split('/')[-1]

        # Создаем документ Excel
        wb = openpyxl.Workbook()
        # Переименовываем лист чтобы в итоговом файле не было пустого листа
        ren_sheet = wb['Sheet']
        ren_sheet.title = 'Сводная таблица'
        # Создаем заголовки для сводной таблицы
        wb['Сводная таблица']['A1'] = 'Наименование ПОО'
        wb['Сводная таблица']['B1'] = 'Наименование государственной услуги'
        wb['Сводная таблица']['C1'] = 'Категория потребителей государственной услуги'
        wb['Сводная таблица']['D1'] = 'Уникальный номер реестровой записи'
        wb['Сводная таблица']['E1'] = 'Профессии и специальности по программам среднего профессионального образования'
        wb['Сводная таблица']['F1'] = 'Категория потребителей'
        wb['Сводная таблица']['G1'] = 'Формы образования и формы реализации образовательных программ'
        wb['Сводная таблица']['H1'] = 'Госзадание №1'
        wb['Сводная таблица']['I1'] = 'Госзадание №2'
        wb['Сводная таблица']['J1'] = 'исполнено на отчетную дату'
        wb['Сводная таблица']['K1'] = 'причина отклонения'
        wb['Сводная таблица']['L1'] = '5% Допустимое отклонение в единицах'
        wb['Сводная таблица']['M1'] = 'Реальное отклонение в единицах'
        wb['Сводная таблица']['N1'] = 'Отклонение в процентах'
        wb['Сводная таблица']['O1'] = 'Отклонение больше 5%'

        wb['Сводная таблица'].column_dimensions['B'].width = 50
        wb['Сводная таблица'].column_dimensions['C'].width = 50
        wb['Сводная таблица'].column_dimensions['D'].width = 30
        wb['Сводная таблица'].column_dimensions['E'].width = 70
        wb['Сводная таблица'].column_dimensions['F'].width = 50
        wb['Сводная таблица'].column_dimensions['G'].width = 15
        wb['Сводная таблица'].column_dimensions['H'].width = 15
        wb['Сводная таблица'].column_dimensions['I'].width = 15
        wb['Сводная таблица'].column_dimensions['J'].width = 15
        wb['Сводная таблица'].column_dimensions['K'].width = 60
        wb['Сводная таблица'].column_dimensions['L'].width = 30
        wb['Сводная таблица'].column_dimensions['M'].width = 30

        # Создаем заливку которой будет отмечать ячейки где отклонение больше 5 %
        redFill = PatternFill(start_color='960018',
                              end_color='960018',
                              fill_type='solid')

        check_name_file = ''

        try:
            check_name_file = f'{name_poo} Госзадание №1'
            print(f'{path_to_data}/Госзадание №1.xlsx')

            df_1 = pd.read_excel(f'{path_to_data}/{name_poo} Госзадание №1.xlsx')
            # Отбираем колонки чтобы сделать из них мультииндекс
            ml_1 = df_1[['Наименование государственной услуги', 'Категория потребителей государственной услуги',
                         'Уникальный номер реестровой записи',
                         'Профессии и специальности по программам среднего профессионального образования']]

            # Отбираем колонки с данными
            data_1 = df_1[
                ['Категория потребителей', 'Формы образования и формы реализации образовательных программ',
                 'Госзадание №1']]

            # Создаем мультиндекс
            ml_df1 = pd.MultiIndex.from_frame(ml_1, names=['Наименование государственной услуги',
                                                           'Категория потребителей государственной услуги',
                                                           'Уникальный номер реестровой записи',
                                                           'Профессии и специальности по программам среднего профессионального образования'])

            # Присваиваем мультиндекс
            data_1.index = ml_df1

            # Открываем госзадание №2
            check_name_file = f'{name_poo} Госзадание №2'
            df_2 = pd.read_excel(f'{path_to_data}/{name_poo} Госзадание №2.xlsx')

            # Создаем мультииндекс
            ml_2 = df_2[['Наименование государственной услуги', 'Категория потребителей государственной услуги',
                         'Уникальный номер реестровой записи',
                         'Профессии и специальности по программам среднего профессионального образования']]

            # Забираем данные которые нам нужны
            data_2 = df_2[['Госзадание №2']]

            ml_df2 = pd.MultiIndex.from_frame(ml_2, names=['Наименование государственной услуги',
                                                           'Категория потребителей государственной услуги',
                                                           'Уникальный номер реестровой записи',
                                                           'Профессии и специальности по программам среднего профессионального образования'])

            data_2.index = ml_df2

            gos_zad_df = pd.concat([data_1, data_2], axis=1, join='outer')

            # Открываем файл отчета
            check_name_file = f'{name_poo} Отчет'
            gos_report = pd.read_excel(f'{path_to_data}/{name_poo} Госзадание Отчет.xlsx')

            # Создаем мультииндекс
            ml_report_df = gos_report[
                ['Наименование государственной услуги', 'Категория потребителей государственной услуги',
                 'Уникальный номер реестровой записи',
                 'Профессии и специальности по программам среднего профессионального образования']]

            data_report = gos_report[['исполнено на отчетную дату', 'причина отклонения']]

            ml_report = pd.MultiIndex.from_frame(ml_report_df, names=['Наименование государственной услуги',
                                                                      'Категория потребителей государственной услуги',
                                                                      'Уникальный номер реестровой записи',
                                                                      'Профессии и специальности по программам среднего профессионального образования'])

            data_report.index = ml_report

            # Соединяем в итоговый датафрейм
            itog_df = pd.concat([gos_zad_df, data_report], axis=1, join='outer')

            # Добавляем нужные столбы
            itog_df['5% Допустимое отклонение в единицах'] = (itog_df['Госзадание №2'] / 100 * 5).round(0)

            itog_df['Реальное отклонение в единицах'] = (
                    itog_df['Госзадание №2'] - itog_df['исполнено на отчетную дату']).abs()

            itog_df['Отклонение в процентах'] = (
                    itog_df['Реальное отклонение в единицах'] * 100 / itog_df['Госзадание №2']).round(2)

            itog_df['Отклонение больше 5%'] = itog_df['Отклонение в процентах'].apply(
                lambda x: 'Да' if x > 5 else 'Нет')

            itog_df['Отклонение в процентах'] = itog_df['Отклонение в процентах'].astype(str) + '%'

            itog_df = itog_df.reset_index()

            itog_df.insert(0, 'Наименование ПОО', name_poo)
            for r in dataframe_to_rows(itog_df, index=False, header=False):
                if len(r) != 1:
                    wb[f'Сводная таблица'].append(r)

        except FileNotFoundError:
            print(f'В папке {name_poo} Файл {check_name_file} Не найден!!! Проверьте названия файлов внутри папки')
        except KeyError:
            print(f'Проверьте названия колонок в файле {check_name_file}')
        else:
            # # Добавляем перенос
            # for cell in wb['Сводная таблица']['B']:
            #     cell.alignment = Alignment(wrap_text=True)
            # for cell in wb['Сводная таблица']['C']:
            #     cell.alignment = Alignment(wrap_text=True)

            # Выделяем ячейку в зависимости от значения
            for cell in wb['Сводная таблица']['O']:
                if cell.value == 'Да':
                    cell.fill = redFill

            # Получаем текущее время для того чтобы использовать в названии
            t = time.localtime()
            current_time = time.strftime('%H_%M_%S', t)
            # Сохраняем итоговый файл
            wb.save(f'{path_to_end_folder}/Сводный отчет по выполнению объемов госзадания {current_time}.xlsx')
            messagebox.showinfo('ЦОПП Бурятия','Обработка успешно завершена!!!')

if __name__=='__main__':
    window = Tk()
    window.title('ЦОПП Бурятия')
    window.geometry('700x860')
    window.resizable(False, False)


    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку подсчета бюджетных заявок
    tab_processing_scope_state_task = ttk.Frame(tab_control)
    tab_control.add(tab_processing_scope_state_task, text='Обработка объема услуг')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку Обработка объема услуг
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_processing_scope_state_task,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nОбработка показателей объема образовательных госуслуг')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img_scope_state_task = resource_path('logo.png')
    img_scope_state_task = PhotoImage(file=path_to_img_scope_state_task)
    Label(tab_processing_scope_state_task,
          image=img_scope_state_task
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_scope_state_task = LabelFrame(tab_processing_scope_state_task, text='Подготовка')
    frame_data_for_scope_state_task.grid(column=0, row=2, padx=10)


    # Создаем кнопку Выбрать папку где находятся данные
    btn_choose_folder_data = Button(frame_data_for_scope_state_task, text='1) Выберите папку с данными', font=('Arial Bold', 20),
                                       command=select_folder_data)
    btn_choose_folder_data.grid(column=0, row=3, padx=10, pady=10)
    #
    # Создаем кнопку для выбора папки куда будет генерироваться файл

    btn_choose_end_folder_data = Button(frame_data_for_scope_state_task, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                       command=select_end_folder
                                       )
    btn_choose_end_folder_data.grid(column=0, row=4, padx=10, pady=10)

    # Создаем переменную хранящую тип документа, в зависимости от значения будет использоваться та или иная функция
    group_rb_type_doc = IntVar()
    # Создаем фрейм для размещения переключателей(pack и грид не используются в одном контейнере)
    frame_rb_type_doc = LabelFrame(tab_processing_scope_state_task, text='Выберите режим обработки')
    frame_rb_type_doc.grid(column=0, row=5, padx=10)
    #
    Radiobutton(frame_rb_type_doc, text='Для Министерства образования и науки', variable=group_rb_type_doc, value=0).pack()
    Radiobutton(frame_rb_type_doc, text='Для ПОО', variable=group_rb_type_doc, value=1).pack()

    # Создаем кнопку для создания документов из таблиц с произвольной структурой
    btn_processing_scope_state_task = Button(tab_processing_scope_state_task, text='Обработать',
                                             font=('Arial Bold', 20),
                                             command=proccessing_scope_state_task
                                             )
    btn_processing_scope_state_task.grid(column=0, row=6, padx=10, pady=10)
    window.mainloop()